from flask import Flask, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import os
import logging
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from flask import request


app = Flask(__name__)
CORS(app)

# Load full pollution data
def load_pollution_data():
    try:
        return pd.read_csv('cleaned_pollution_data.csv')
    except FileNotFoundError:
        print("Error: 'cleaned_pollution_data.csv' not found.")
        return pd.DataFrame()
    except Exception as e:
        print(f"Error loading data: {e}")
        return pd.DataFrame()

# Generate bar chart for PDF
def create_bar_chart(data):
    plt.figure(figsize=(10, 5))
    plt.bar(data['city'], data['pm25'], color='#1f77b4')
    plt.xlabel('City')
    plt.ylabel('Average PM2.5 (µg/m³)')
    plt.title('Average PM2.5 Levels by City')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    chart_path = 'pm25_chart.png'
    plt.savefig(chart_path)
    plt.close()
    return chart_path

# Create PDF report
def create_pdf_report(filename, data):
    try:
        doc = SimpleDocTemplate(filename, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Pollution Monitoring Report", styles['Title']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph("Average PM2.5 Levels by City", styles['Heading2']))
        elements.append(Spacer(1, 12))

        table_data = [data.columns.tolist()] + data.values.tolist()
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 12))

        chart_path = create_bar_chart(data)
        elements.append(Paragraph("PM2.5 Levels Visualization", styles['Heading2']))
        elements.append(Spacer(1, 12))
        elements.append(Image(chart_path, width=400, height=200))

        doc.build(elements)
        os.remove(chart_path)
        return True
    except Exception as e:
        print(f"Error generating PDF: {e}")
        return False

# Create Excel report
def create_excel_report(filename, data):
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            data.to_excel(writer, sheet_name='PM2.5 Report', index=False)
            workbook = writer.book
            sheet = workbook['PM2.5 Report']
            header_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            cell_align = Alignment(horizontal='center', vertical='center')

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = cell_align
                cell.border = cell_border

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
                for cell in row:
                    cell.alignment = cell_align
                    cell.border = cell_border
                    if cell.column_letter == 'B':
                        cell.number_format = '0.00'

            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                sheet.column_dimensions[column].width = adjusted_width
        return True
    except Exception as e:
        print(f"Error generating Excel: {e}")
        return False

# Home route
@app.route('/', methods=['GET'])
def home():
    return "Welcome to the Pollution Data API! Try /api/daily_pollution, /api/avg_by_city, /api/generate_pdf, or /api/generate_excel."

# Date-wise full pollution data
@app.route('/api/daily_pollution', methods=['GET'])
def get_daily_pollution():
    df = load_pollution_data()
    if df.empty:
        return jsonify({"error": "No data available"}), 500
    daily_data = df.groupby('date')[['pm25', 'pm10', 'no2', 'o3']].mean().reset_index()
    return jsonify(daily_data.to_dict(orient='records'))

# City-wise average data
@app.route('/api/avg_by_city', methods=['GET'])
def get_avg_by_city():
    df = load_pollution_data()
    if df.empty:
        return jsonify({"error": "No data available"}), 500
    avg_by_city = df.groupby('city')[['pm25', 'pm10', 'no2', 'o3']].mean().reset_index()
    return jsonify(avg_by_city.to_dict(orient='records'))

# Generate PDF report endpoint
@app.route('/api/generate_pdf', methods=['GET'])
def generate_pdf():
    df = load_pollution_data()
    if df.empty:
        return jsonify({"error": "No data available for PDF report"}), 500
    avg_pm25 = df.groupby('city')['pm25'].mean().reset_index()
    filename = 'pollution_report.pdf'
    if create_pdf_report(filename, avg_pm25):
        return send_file(filename, as_attachment=True)
    return jsonify({"error": "Failed to generate PDF report"}), 500

# Generate Excel report endpoint
@app.route('/api/generate_excel', methods=['GET'])
def generate_excel():
    df = load_pollution_data()
    if df.empty:
        return jsonify({"error": "No data available for Excel report"}), 500
    avg_pm25 = df.groupby('city')['pm25'].mean().reset_index()
    filename = 'pollution_report.xlsx'
    if create_excel_report(filename, avg_pm25):
        return send_file(filename, as_attachment=True)
    return jsonify({"error": "Failed to generate Excel report"}), 500



logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

if __name__ == '__main__':
    app.run(debug=True, port=5000)