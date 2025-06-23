import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Load data
try:
    df = pd.read_csv('cleaned_pollution_data.csv')
    avg_pm25 = df.groupby('city')['pm25'].mean().reset_index()
except FileNotFoundError:
    print("Error: 'cleaned_pollution_data.csv' not found.")
    avg_pm25 = pd.DataFrame(columns=['city', 'pm25'])

# Create Excel report
def create_excel_report(filename, data):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        data.to_excel(writer, sheet_name='PM2.5 Report', index=False)

        # Access the openpyxl workbook and sheet
        workbook = writer.book
        sheet = workbook['PM2.5 Report']

        # Define styles
        header_fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        cell_align = Alignment(horizontal='center', vertical='center')

        # Style the header
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = cell_align
            cell.border = cell_border

        # Style the data cells
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = cell_align
                cell.border = cell_border
                if cell.column_letter == 'B':  # PM2.5 column
                    cell.number_format = '0.00'

        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width
    # No explicit save() needed; the 'with' statement handles it

if __name__ == '__main__':
    create_excel_report('pollution_report.xlsx', avg_pm25)